import frappe
import os
import csv
import io
import base64
import mimetypes


def parse_files(file_urls: list) -> dict:
    """Parse uploaded files into text content and/or base64 images.

    Args:
        file_urls: List of Frappe file URLs (e.g., ["/files/receipt.pdf"])

    Returns:
        {"text": "extracted text...", "images": [{"media_type": "image/png", "data": "base64..."}]}
    """
    text_parts = []
    images = []

    for url in file_urls:
        try:
            file_path = _get_file_path(url)
            if not file_path or not os.path.exists(file_path):
                text_parts.append(f"[Could not read file: {url}]")
                continue

            ext = os.path.splitext(file_path)[1].lower()

            if ext == ".pdf":
                text_parts.append(_parse_pdf(file_path))
            elif ext in (".xlsx", ".xls"):
                text_parts.append(_parse_excel(file_path))
            elif ext == ".csv":
                text_parts.append(_parse_csv(file_path))
            elif ext in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
                img_data = _parse_image(file_path)
                if img_data:
                    images.append(img_data)
            elif ext in (".txt", ".md", ".json", ".log"):
                with open(file_path, "r", errors="replace") as f:
                    text_parts.append(f.read()[:50000])  # Limit text size
            else:
                text_parts.append(f"[Unsupported file type: {ext}]")

        except Exception as e:
            text_parts.append(f"[Error reading {url}: {str(e)}]")

    return {
        "text": "\n\n".join(text_parts) if text_parts else "",
        "images": images,
    }


def _get_file_path(url: str) -> str:
    """Convert Frappe file URL to filesystem path."""
    if url.startswith("/files/"):
        return frappe.get_site_path("public", "files", url.replace("/files/", ""))
    elif url.startswith("/private/files/"):
        return frappe.get_site_path("private", "files", url.replace("/private/files/", ""))

    # Try getting from File doctype
    file_doc = frappe.db.get_value("File", {"file_url": url}, "name")
    if file_doc:
        f = frappe.get_doc("File", file_doc)
        return f.get_full_path()

    return None


def _parse_pdf(file_path: str) -> str:
    """Extract text from PDF using PyPDF2."""
    try:
        from PyPDF2 import PdfReader
    except ImportError:
        return "[PyPDF2 not installed — cannot parse PDF]"

    reader = PdfReader(file_path)
    text_parts = []
    for i, page in enumerate(reader.pages[:50]):  # Limit to 50 pages
        text = page.extract_text()
        if text:
            text_parts.append(f"--- Page {i + 1} ---\n{text}")

    return "\n\n".join(text_parts) if text_parts else "[No text extracted from PDF]"


def _parse_excel(file_path: str) -> str:
    """Extract data from Excel file using openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return "[openpyxl not installed — cannot parse Excel]"

    wb = load_workbook(file_path, read_only=True, data_only=True)
    text_parts = []

    for sheet_name in wb.sheetnames[:5]:  # Limit to 5 sheets
        sheet = wb[sheet_name]
        text_parts.append(f"--- Sheet: {sheet_name} ---")

        rows = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i >= 200:  # Limit rows
                text_parts.append(f"... ({sheet.max_row - 200} more rows)")
                break
            row_values = [str(cell) if cell is not None else "" for cell in row]
            rows.append(" | ".join(row_values))

        text_parts.append("\n".join(rows))

    wb.close()
    return "\n\n".join(text_parts)


def _parse_csv(file_path: str) -> str:
    """Extract data from CSV file."""
    rows = []
    with open(file_path, "r", errors="replace") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i >= 200:
                rows.append("... (more rows)")
                break
            rows.append(" | ".join(row))

    return "\n".join(rows) if rows else "[Empty CSV]"


def _parse_image(file_path: str) -> dict:
    """Read image file as base64 for Claude vision."""
    mime_type = mimetypes.guess_type(file_path)[0] or "image/png"

    with open(file_path, "rb") as f:
        data = base64.b64encode(f.read()).decode("utf-8")

    return {
        "media_type": mime_type,
        "data": data,
    }
